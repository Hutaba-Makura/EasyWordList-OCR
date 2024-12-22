import openpyxl as px

DATA_PATH = 'data.xlsx'
wb = px.Workbook() # ワークブックを作成
ws = wb.active # ワークシートを取得

# data.xlsxに単語と意味を追加する関数
def addWord(wordset: dict):
    # wordset の各単語と意味を追加
    for word, mean in wordset.items():
        ws.append([word, mean])
    
    wb.save(DATA_PATH)  # 保存

# data.xlsxから単語と意味を取得する関数
def getWords():
    words = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
        words[row[0]] = row[1]
    
    return words

# 送られたデータから特定の単語を編集する関数
# 行番号を取得して、その行の意味を編集する
def editWord(wordset: dict, row):
    # 指定された行の単語を削除
    ws.delete_rows(row)
    # 指定された行に新しい単語を挿入
    ws.insert_rows(row)
    ws.cell(row=row, column=1, value=list(wordset.keys())[0])
    ws.cell(row=row, column=2, value=list(wordset.values())[0])

    wb.save(DATA_PATH)  # 保存

# 送られたデータから特定の単語を削除する関数
# 行番号を取得して、その行を削除する
def deleteWord(row):
    ws.delete_rows(row)

    wb.save(DATA_PATH)  # 保存

# 受け取った辞書からWordlistを一新する関数
def updateWordlist(wordlist):
    ws.delete_rows(2, ws.max_row) # 既存のデータを削除

    for word, mean in wordlist.items():
        ws.append([word, mean])
    
    wb.save(DATA_PATH)  # 保存